import os
from openpyxl import load_workbook
from openpyxl import Workbook
import re
import window_alert


class Turnus:
    def __init__(self, number):
        self.number = number
        self.name = 'name'
        self.surname = 'surname'
        self.groups = 'groups'
        self.night_groups = 'night_groups'
        self.gender = 'gender'

        self.kids_names = []
        self.groups_mens = 1
        self.groups_women = 2
        self.counter_mens = 0
        self.counter_women = 0
        self.counter_night_group = 0
        self.wb = Turnus.create_workbook(number)
        self.sheet = self.wb.active

    def read(self):
        wb = load_workbook(f'files/plik{self.number}.xlsx')
        sheet_name = "DZIECI"
        if sheet_name in wb.sheetnames.pop(0).upper():

            ws = wb.worksheets[0]

            for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    if cell.value is None:
                        self.wb.save(f"files/turnus{self.number}.xlsx")
                        break
                    elif cell.value in self.kids_names:
                        continue
                    else:
                        self.kids_names.append(cell.value)
                        self.name, self.surname = re.split(r'\s(?=[a-z][A-Z])|\s', cell.value, maxsplit=1)
                        Turnus.gender_name_create(self)

        else:
            window_alert.window_alert('Sprawdz poprawnosc tego pliku gdyz nie zawiera on arkuszu dzieci')
            os.remove(f'files/turnus{self.number}.xlsx')
        os.remove(f'files/plik{self.number}.xlsx')

    def gender_name_create(self):
        names_women = ['NIKOL', 'NICOLE', 'NIKOLE', 'NEL', 'BERNEIKE', 'DORIS', 'DOLORES', 'NELLY', 'SALOME']

        if str(self.name[-1]).upper() != 'A' or (str(self.name).upper() == 'KUBA' and str(self.name).upper()
                                                 not in names_women):
            self.counter_mens += 1
            self.counter_night_group += 1
            self.gender = 'M'
        else:
            self.counter_women += 1
            self.counter_night_group += 1
            self.gender = 'F'

        Turnus.allocation_groups(self)

    def allocation_groups(self):
        if self.gender == 'M':
            if self.counter_mens / 11 % 1 == 0:
                self.groups_mens += 2
            self.groups = self.groups_mens
        else:
            if self.counter_women / 11 % 1 == 0:
                self.groups_women += 2
            self.groups = self.groups_women

        self.night_groups = int(self.counter_night_group / 15) + 1

        Turnus.save(self)

    def save(self):
        sheet = self.wb.active
        row_data = [self.name, self.surname, self.groups, self.night_groups, self.gender]
        sheet.append(row_data)

    @staticmethod
    def create_workbook(number):
        wb = Workbook()

        ws = wb.active

        column_data = ['Name', 'Surname', 'Groups', 'NightGroups', 'Gender']
        ws.append(column_data)

        wb.save(f"files/turnus{number}.xlsx")

        return wb

